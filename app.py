import os
from distutils.log import debug
from fileinput import filename
import shutil
import openpyxl
from openpyxl.utils import column_index_from_string
from flask import *
from datetime import datetime
from werkzeug.utils import secure_filename
app = Flask(__name__)

UPLOAD_FOLDER = "D:\\Barang Given\\Fiverr\\Nikolay\\bid-app-flask\\uploads"
ORIGINAL_PATH =  "D:\\Barang Given\\Fiverr\\Nikolay\\bid-app-flask"
ALLOWED_EXTENSIONS = {'xls','xlsx'}

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['ORIGINAL_PATH'] = ORIGINAL_PATH
@app.route('/')
def main():
	return render_template("index.html")

@app.route('/success', methods = ['POST'])
def success():
	if request.method == 'POST':
		file60days = request.files['60-days-file']
		fileYesterday = request.files['yesterday-file']
		if file60days.filename == '' and fileYesterday.filename == '':
			flash('No selected file')
			return redirect(request.url)
		if (file60days and allowed_file(file60days.filename)) and (fileYesterday and allowed_file(fileYesterday.filename)):
			now = datetime.now()
			filename60days =   secure_filename(file60days.filename)
			filenameYesterday = secure_filename(fileYesterday.filename)
			stringDate = now.strftime("%m-%d-%Y %H-%M-%S") 
			savepath = os.path.join(app.config['UPLOAD_FOLDER'], stringDate)
			os.mkdir(savepath)
			file60days.save(os.path.join(savepath, filename60days))
			fileYesterday.save(os.path.join(savepath, filenameYesterday))

			return "Upload Succesfully"


def manual(PathFile60days, PathfileYesterday) : 
	fileYesterday = xw.books.open(PathfileYesterday)
	file60 = xw.books.open(PathFile60days)
	final = xw.Book()
	final.save('final.xlsx')
	lrow1 = fileYesterday.Cells(Rows.count, 1).End(xlUp).Row
	final.Cells.Clear

	# Load the excel workbook
	workbook = openpyxl.load_workbook('example.xlsx')
	sht1 = workbook['Sheet1']
	sht2 = workbook['Sheet2']
	sht3 = workbook['Sheet3']
	sht4 = workbook['Sheet4']

	# Find last row of column 1 in sht1
	lrow1 = sht1.max_row
	while sht1.cell(row=lrow1, column=1).value is None:
		lrow1 -= 1

	# Clear data in sht3
	sht3.delete_rows(1, sht3.max_row)

	# Copy the first row from sht1 to sht3
	for col in range(1, sht1.max_column + 1):
		sht3.cell(row=1, column=col).value = sht1.cell(row=1, column=col).value

	# Loop through each row in sht1 and apply the necessary conditions
	k = 2
	for i in range(2, lrow1 + 1):
		if sht1.cell(row=i, column=26).value == "":
			sht1.cell(row=i, column=26).value = sht1.cell(row=i, column=25).value

		if sht1.cell(row=i, column=34).value > sht4.cell(row=11, column=8).value:
			if sht1.cell(row=i, column=42).value < (sht4.cell(row=17, column=10).value * 0.01) or sht1.cell(row=i, column=42).value > (sht4.cell(row=17, column=11).value * 0.01):
				if sht4.cell(row=19, column=10).value.upper() == "NO" or (sht4.cell(row=19, column=10).value.upper() == "YES" and sht1.cell(row=i, column=35).value > 0):
					sht1.cell(row=i, column=53).value = f"=MATCH(H{i},'60 Days'!H:H,0)"
					if not isinstance(sht1.cell(row=i, column=53).value, str):
						if sht2.cell(row=sht1.cell(row=i, column=53).value, column=42).value < (sht4.cell(row=18, column=10).value * 0.01) or sht2.cell(row=sht1.cell(row=i, column=53).value, column=42).value > (sht4.cell(row=18, column=11).value * 0.01):
							if sht2.cell(row=sht1.cell(row=i, column=53).value, column=35).value > sht4.cell(row=16, column=10).value:
								for col in range(1, sht1.max_column + 1):
									sht3.cell(row=k, column=col).value = sht1.cell(row=i, column=col).value
								sht3.cell(row=k, column=52).value = sht2.cell(row=sht1.cell(row=i, column=53).value, column=42).value
								impres = sht3.cell(row=k, column=34).value
								if sht4.cell(row=22, column=5).value == "BY Percentage":
									# ans = worksheet_function.lookup(impres, sht4.range('C11:C
									return "cool"



def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

if __name__ == '__main__':
	app.run(debug=True)



# from flask import Flask, redirect, url_for, render_template, request
# from datetime import datetime


# app =  Flask(__name__)

# @app.route("/")
# def home():
#     totalData = len(readData())
#     Pangkat = []
#     Faktorial = []
#     for i, data in enumerate(readData()):
#       data_break = data.split(',');
#       if data_break[1] == "Perpangkatan":
#          Pangkat.append(data_break)
#       else:
#          Faktorial.append(data_break)
    
#     persenFaktorial = len(Faktorial)/totalData * 100
#     persenPangkat = len(Pangkat)/totalData * 100
#     return render_template("index.html", totalPangkat = len(Pangkat), totalFaktorial = len(Faktorial), persenFaktorial = persenFaktorial, persenPangkat = persenPangkat)

# # @app.route("/menu")
# # def menu():
# #     return render_template("menu.html")


# # @app.route("/faktorial", methods=['POST', 'GET'])
# # def faktorial():
# #      if request.method == 'POST':
# #         now = datetime.now()
# #         tanggal = now.strftime("%m/%d/%Y : %H:%M:%S")
# #         jenis = "Faktorial"
# #         angka = int(request.form['faktorial'])
# #         hasil = faktorialkan(angka)
# #         with open("data.txt", "a") as file:
# #          file.write(f"\n{tanggal},{jenis},{angka},{hasil}")
# #         return render_template('faktorial.html', hasil = hasil)


# #      else:
# #         return render_template('faktorial.html')
   

# # @app.route("/pangkat", methods=['POST', 'GET'])
# # def pangkat():
# #      if request.method == 'POST':
# #         now = datetime.now()
# #         tanggal = now.strftime("%m/%d/%Y : %H:%M:%S")
# #         jenis = "Perpangkatan"
# #         angka = int(request.form['angka'])
# #         pangkat = int(request.form['pangkat'])
# #         hasil = perpangkatan(angka, pangkat)
# #         with open("data.txt", "a") as file:
# #          file.write(f"\n{tanggal},{jenis},{angka},{pangkat},{hasil}")
# #         return render_template('pangkat.html', hasil = hasil)

# #      else:
# #         return render_template('pangkat.html')
   

# # @app.route("/menu")
# # def menu():
# #     return render_template("menu.html")

# # @app.route("/history-pangkat")
# # def history_pangkat():
# #     dataAll = []
# #     for i, data in enumerate(readData()):
# #       data_break = data.split(',');
# #       if data_break[1] == "Perpangkatan":
# #          dataAll.append(data_break)
# #    #          data['tanggal'] = data_break[0]
# #    #          data['jenis'] = data_break[1]
# #    #          data['angka'] = data_break[2]
# #    #          data['pangkat'] = data_break[3]
# #    #          data['hasil'] = data_break[4]
# #     return render_template("history-pangkat.html", data=dataAll)


# # @app.route("/history-faktorial")
# # def history_faktorial():
# #     dataAll=[]
# #     for i, data in enumerate(readData()):
# #       data_break = data.split(',')
      
# #       if data_break[1] == "Faktorial":
# #             dataAll.append(data_break)
# #             # dataAll[i]['tanggal'] = data_break[0]
# #             # dataAll[i]['jenis'] = data_break[1]
# #             # dataAll[i]['angka'] = data_break[2]
# #             # dataAll[i]['hasil'] = data_break[3]
# #     return render_template("history-faktorial.html",data=dataAll)


# if __name__ == "__main__":
#     app.run(debug=True)